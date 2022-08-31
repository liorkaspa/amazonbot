import os
import bs4
from bs4 import BeautifulSoup
import requests
import pandas as pd
import numpy as np
from twilio.rest import Client
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
#imports

# Find your Account SID and Auth Token at twilio.com/console
# and set the environment variables. See http://twil.io/secure
phone_no_to = '+972506468199'


def send_sms(phone_no_to, price):
    account_sid = os.environ["account_sid"]
    auth_token = os.environ["auth_token"]
    client =  Client(account_sid, auth_token)

    message = client.messages \
        .create(
        body="Layerd protein bar price is %d, Consider to order now using the following link: %s" % (
        price, "https://www.myprotein.co.il/sports-nutrition/layered-protein-bar/12116519.html?affil=thggpsad&switchcurrency=ILS&shippingcountry=IL&variation=12385227&thg_ppc_campaign=71700000083245846&product_id=12385227&gclid=Cj0KCQiAjc2QBhDgARIsAMc3SqTcG4xAsiw8XYyh9433JntHDhhhZdNQDzrcxjFspUhRngRp41hCZjsaAr4_EALw_wcB&gclsrc=aw.ds"),
        from_='+18596462720',
        to=phone_no_to
    )


headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.102 Safari/537.36',
    'Accept-Language': 'en-US,en;q=0.9,he;q=0.8'}

res = requests.get(url="https://www.myprotein.co.il/sports-nutrition/layered-protein-bar/12116519.html",
                   headers=headers)
soup = BeautifulSoup(res.text, 'html.parser')  # parse as html
price_float = float(soup.find("span", class_="productPrice_schema productPrice_priceAmount").text)
price = int(np.round(price_float, 0))

def to_excel():
    try:
        date_changed = True
        book = load_workbook('Layerd Bar Prices.xlsx')

    except:
        df = pd.DataFrame(columns=["Date", "Price"])
        df.to_excel("Layerd Bar Prices.xlsx")
        book = load_workbook('Layerd Bar Prices.xlsx')

    finally:
        ws = book.worksheets[0]
        date = datetime.today().strftime('%Y-%m-%d')
        last_price = ws.cell(row=ws.max_row, column=3).value
        last_date = ws.cell(row=ws.max_row, column=2).value
        if last_date == date:
            date_changed = False
        if date_changed:
            ws.append(["", date, price])
        book.save("Layerd Bar Prices.xlsx")
        return int(last_price), date_changed


date_changed, old_price = to_excel()
if price != old_price and date_changed:
    send_sms(phone_no_to , price)
